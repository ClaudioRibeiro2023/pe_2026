#!/usr/bin/env python3
"""
Validador de Integração: Excel → Supabase
Verifica se os dados dos planos de ação (Excel) estão corretamente refletidos no banco.

Uso:
    python scripts/validate_action_plan_integration.py --area rh
    python scripts/validate_action_plan_integration.py --all
"""

import os
import sys
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass

# Tenta importar openpyxl para leitura do Excel
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("AVISO: openpyxl não instalado. Instale com: pip install openpyxl")

# Tenta importar psycopg2 para conexão direta
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    HAS_PSYCOPG2 = True
except ImportError:
    HAS_PSYCOPG2 = False


@dataclass
class Initiative:
    """Representa uma iniciativa do plano de ação."""
    code: str
    title: str
    type: str
    priority: str
    pillar: str
    subpillar: str
    motor: str
    okr: str
    kr: str
    owner: str
    sponsor: str
    deadline: str
    cost: str
    status: str = "PLANEJADA"
    area: Optional[str] = None


@dataclass
class ValidationResult:
    """Resultado de uma validação."""
    passed: bool
    message: str
    details: Optional[Dict] = None


class ActionPlanValidator:
    """Validador de integração de planos de ação."""
    
    # Mapeamento de áreas para slugs
    AREA_SLUGS = {
        'RH': 'rh',
        'RH / Pessoas': 'rh',
        'Marketing': 'marketing',
        'P&D': 'pd',
        'P&D / Produto / Dados': 'pd',
        'Operação': 'operacoes',
        'CS': 'cs',
        'CS / Relacionamento': 'cs',
        'Comercial': 'comercial',
        'Financeiro': 'financeiro',
    }
    
    # INITs setoriais esperadas por área (do generate_action_plans.py)
    EXPECTED_SECTORIAL_INITS = {
        'rh': [
            'INIT-RH-301', 'INIT-RH-302', 'INIT-RH-303',
            'INIT-RH-304', 'INIT-RH-305', 'INIT-RH-306'
        ],
        'marketing': [
            'INIT-MKT-101', 'INIT-MKT-102', 'INIT-MKT-103',
            'INIT-MKT-104', 'INIT-MKT-105', 'INIT-MKT-106', 'INIT-MKT-107'
        ],
        'pd': [
            'INIT-PD-251', 'INIT-PD-252', 'INIT-PD-253',
            'INIT-PD-254', 'INIT-PD-255', 'INIT-PD-256'
        ],
        'operacoes': [
            'INIT-OP-151', 'INIT-OP-152', 'INIT-OP-153', 'INIT-OP-154',
            'INIT-OP-155', 'INIT-OP-156', 'INIT-OP-157', 'INIT-OP-158'
        ],
        'cs': [
            'INIT-CS-201', 'INIT-CS-202', 'INIT-CS-203', 'INIT-CS-204',
            'INIT-CS-205', 'INIT-CS-206', 'INIT-CS-207', 'INIT-CS-208'
        ],
        'comercial': [
            'INIT-COM-401', 'INIT-COM-402', 'INIT-COM-403',
            'INIT-COM-404', 'INIT-COM-405'
        ],
        'financeiro': [
            'INIT-FIN-351', 'INIT-FIN-352', 'INIT-FIN-353', 'INIT-FIN-354',
            'INIT-FIN-355', 'INIT-FIN-356', 'INIT-FIN-357', 'INIT-FIN-358'
        ],
    }
    
    def __init__(self, excel_dir: str = None):
        """Inicializa o validador."""
        self.excel_dir = excel_dir or self._find_excel_dir()
        self.results: List[ValidationResult] = []
        
    def _find_excel_dir(self) -> str:
        """Encontra o diretório com os arquivos Excel."""
        # Tenta caminhos relativos comuns
        possible_paths = [
            'docs/planos-acao/pa.2026',
            '../docs/planos-acao/pa.2026',
            '../../docs/planos-acao/pa.2026',
            'B:/PE_2026/docs/planos-acao/pa.2026',
        ]
        
        for path in possible_paths:
            if os.path.exists(path):
                return path
                
        # Se não encontrar, usa o caminho padrão
        return 'docs/planos-acao/pa.2026'
    
    def _load_excel_data(self, area_slug: str) -> List[Initiative]:
        """Carrega dados do arquivo Excel da área."""
        if not HAS_OPENPYXL:
            print("ERRO: openpyxl não disponível para leitura do Excel")
            return []
            
        filename_map = {
            'rh': 'PE2026_RH.xlsx',
            'marketing': 'PE2026_Marketing.xlsx',
            'pd': 'PE2026_PD.xlsx',
            'operacoes': 'PE2026_Operacao.xlsx',
            'cs': 'PE2026_CS.xlsx',
            'comercial': 'PE2026_Comercial.xlsx',
            'financeiro': 'PE2026_Financeiro.xlsx',
        }
        
        filename = filename_map.get(area_slug)
        if not filename:
            return []
            
        filepath = os.path.join(self.excel_dir, filename)
        if not os.path.exists(filepath):
            print(f"AVISO: Arquivo não encontrado: {filepath}")
            return []
            
        initiatives = []
        
        try:
            wb = load_workbook(filepath, data_only=True)
            
            # Procura na aba "Plano de Acao" ou similar
            target_sheet = None
            for sheet_name in wb.sheetnames:
                if 'plano' in sheet_name.lower() or 'acao' in sheet_name.lower():
                    target_sheet = wb[sheet_name]
                    break
                    
            if not target_sheet and wb.sheetnames:
                target_sheet = wb[wb.sheetnames[-1]]  # Última aba como fallback
                
            if target_sheet:
                # Extrai dados das linhas (ignora header)
                for row in target_sheet.iter_rows(min_row=5, values_only=True):
                    if row and row[0] and str(row[0]).startswith('INIT'):
                        init = Initiative(
                            code=str(row[0]),
                            title=str(row[4]) if len(row) > 4 else '',
                            type=str(row[5]) if len(row) > 5 else '',
                            priority=str(row[6]) if len(row) > 6 else '',
                            pillar=str(row[2]) if len(row) > 2 else '',
                            subpillar=str(row[3]) if len(row) > 3 else '',
                            motor=str(row[1]) if len(row) > 1 else '',
                            okr=str(row[7]) if len(row) > 7 else '',
                            kr=str(row[8]) if len(row) > 8 else '',
                            owner=str(row[9]) if len(row) > 9 else '',
                            sponsor=str(row[10]) if len(row) > 10 else '',
                            deadline=str(row[11]) if len(row) > 11 else '',
                            cost=str(row[12]) if len(row) > 12 else '',
                            status='PLANEJADA',
                            area=area_slug
                        )
                        initiatives.append(init)
                        
            wb.close()
            
        except Exception as e:
            print(f"ERRO ao ler Excel: {e}")
            
        return initiatives
    
    def validate_sectorial_initiatives(self, area_slug: str) -> ValidationResult:
        """Valida se todas as INITs setoriais esperadas existem no banco."""
        expected = set(self.EXPECTED_SECTORIAL_INITS.get(area_slug, []))
        
        # Aqui você faria uma query no Supabase
        # Por enquanto, simulamos verificando o arquivo seed SQL
        seed_file = 'supabase/seeds/07_rh_action_plan_real_data.sql'
        found_in_seed = set()
        
        try:
            if os.path.exists(seed_file):
                with open(seed_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                    for code in expected:
                        if code in content:
                            found_in_seed.add(code)
        except Exception as e:
            return ValidationResult(
                False, 
                f"Erro ao verificar seed: {e}",
                {'expected': len(expected), 'found': 0}
            )
            
        missing = expected - found_in_seed
        
        if missing:
            return ValidationResult(
                False,
                f"INITs faltantes no seed: {', '.join(missing)}",
                {'expected': len(expected), 'found': len(found_in_seed), 'missing': list(missing)}
            )
            
        return ValidationResult(
            True,
            f"Todas as {len(expected)} INITs setoriais encontradas no seed",
            {'expected': len(expected), 'found': len(found_in_seed)}
        )
    
    def validate_excel_data_integrity(self, area_slug: str) -> ValidationResult:
        """Valida integridade dos dados do Excel."""
        initiatives = self._load_excel_data(area_slug)
        
        if not initiatives:
            return ValidationResult(
                False,
                f"Nenhuma iniciativa encontrada no Excel para {area_slug}",
                {'file': f'{area_slug}.xlsx'}
            )
            
        errors = []
        
        for init in initiatives:
            # Validações básicas
            if not init.code.startswith('INIT'):
                errors.append(f"Código inválido: {init.code}")
                
            if init.priority not in ['P0', 'P1', 'P2']:
                errors.append(f"Prioridade inválida em {init.code}: {init.priority}")
                
            if not init.pillar.startswith('P'):
                errors.append(f"Pilar inválido em {init.code}: {init.pillar}")
                
            if len(init.title) < 5:
                errors.append(f"Título muito curto em {init.code}")
                
        if errors:
            return ValidationResult(
                False,
                f"Erros de integridade encontrados: {len(errors)}",
                {'errors': errors[:5], 'total_inits': len(initiatives)}
            )
            
        return ValidationResult(
            True,
            f"Dados do Excel válidos: {len(initiatives)} iniciativas",
            {'total_inits': len(initiatives)}
        )
    
    def validate_budget_consistency(self, area_slug: str) -> ValidationResult:
        """Valida consistência de orçamento."""
        initiatives = self._load_excel_data(area_slug)
        
        total_budget = 0
        budget_errors = []
        
        for init in initiatives:
            try:
                # Tenta extrair valor numérico do campo de custo
                cost_str = init.cost.replace('R$', '').replace(' ', '').strip()
                
                # Remove sufixos como "/ano", "/mês", etc.
                if '/' in cost_str:
                    cost_str = cost_str.split('/')[0]
                    
                # Tenta converter
                if 'K' in cost_str.upper():
                    value = float(cost_str.upper().replace('K', '')) * 1000
                elif 'M' in cost_str.upper():
                    value = float(cost_str.upper().replace('M', '')) * 1000000
                else:
                    # Faixa de valores (ex: "5-10K")
                    if '-' in cost_str:
                        parts = cost_str.replace('K', '').split('-')
                        if len(parts) == 2:
                            value = (float(parts[0]) + float(parts[1])) / 2 * 1000
                        else:
                            value = 0
                    else:
                        value = float(cost_str)
                        
                total_budget += value
                
            except (ValueError, AttributeError) as e:
                budget_errors.append(f"Erro no custo de {init.code}: {init.cost}")
                
        # Verifica se está dentro de limites razoáveis
        if total_budget > 10000000:  # R$ 10M
            return ValidationResult(
                False,
                f"Orçamento total ({total_budget:,.0f}) parece excessivo",
                {'total_budget': total_budget, 'init_count': len(initiatives)}
            )
            
        return ValidationResult(
            True,
            f"Orçamento calculado: R$ {total_budget:,.0f} para {len(initiatives)} iniciativas",
            {'total_budget': total_budget, 'init_count': len(initiatives), 'errors': budget_errors}
        )
    
    def run_all_validations(self, area_slug: str) -> List[ValidationResult]:
        """Executa todas as validações para uma área."""
        results = []
        
        print(f"\n{'='*60}")
        print(f"VALIDANDO PLANO DE AÇÃO: {area_slug.upper()}")
        print(f"{'='*60}")
        
        # Validação 1: INITs setoriais no seed
        result = self.validate_sectorial_initiatives(area_slug)
        results.append(result)
        status = "✅ PASSOU" if result.passed else "❌ FALHOU"
        print(f"\n{status} - INITs Setoriais: {result.message}")
        
        # Validação 2: Integridade do Excel
        result = self.validate_excel_data_integrity(area_slug)
        results.append(result)
        status = "✅ PASSOU" if result.passed else "❌ FALHOU"
        print(f"{status} - Integridade Excel: {result.message}")
        
        # Validação 3: Consistência de orçamento
        result = self.validate_budget_consistency(area_slug)
        results.append(result)
        status = "✅ PASSOU" if result.passed else "⚠️ AVISO"
        print(f"{status} - Orçamento: {result.message}")
        
        return results
    
    def generate_report(self, results: List[ValidationResult]) -> str:
        """Gera relatório de validação."""
        passed = sum(1 for r in results if r.passed)
        total = len(results)
        
        report = []
        report.append("\n" + "="*60)
        report.append("RELATÓRIO DE VALIDAÇÃO")
        report.append("="*60)
        report.append(f"Total de testes: {total}")
        report.append(f"Passaram: {passed}")
        report.append(f"Falharam: {total - passed}")
        report.append(f"Taxa de sucesso: {passed/total*100:.1f}%")
        report.append("="*60)
        
        return "\n".join(report)


def main():
    parser = argparse.ArgumentParser(
        description='Valida integração de planos de ação (Excel → Supabase)'
    )
    parser.add_argument(
        '--area',
        choices=['rh', 'marketing', 'pd', 'operacoes', 'cs', 'comercial', 'financeiro', 'all'],
        default='rh',
        help='Área a validar (padrão: rh)'
    )
    parser.add_argument(
        '--excel-dir',
        help='Diretório com os arquivos Excel (padrão: docs/planos-acao/pa.2026)'
    )
    
    args = parser.parse_args()
    
    validator = ActionPlanValidator(excel_dir=args.excel_dir)
    
    areas_to_validate = []
    if args.area == 'all':
        areas_to_validate = list(validator.EXPECTED_SECTORIAL_INITS.keys())
    else:
        areas_to_validate = [args.area]
        
    all_results = []
    for area in areas_to_validate:
        results = validator.run_all_validations(area)
        all_results.extend(results)
        
    # Relatório final
    print(validator.generate_report(all_results))
    
    # Exit code baseado no resultado
    failed = sum(1 for r in all_results if not r.passed)
    sys.exit(0 if failed == 0 else 1)


if __name__ == '__main__':
    main()
