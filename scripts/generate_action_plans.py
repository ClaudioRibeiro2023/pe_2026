#!/usr/bin/env python3
"""
Gerador de Planos de Ação Setoriais PE2026
Gera 8 workbooks Excel: 7 por área + 1 corporativo consolidado.
Dados canônicos extraídos dos DOCs 04, 06, 08, 09.
"""

import os
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ─────────────────────────────────────────────────────────
# CONSTANTES DE ESTILO
# ─────────────────────────────────────────────────────────

FONT_BODY = Font(name="Calibri", size=11)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Calibri", size=14, bold=True)
FONT_SUBTITLE = Font(name="Calibri", size=12, bold=True)
FONT_DASH_BIG = Font(name="Calibri", size=28, bold=True, color="1F3864")
FONT_DASH_LABEL = Font(name="Calibri", size=11, color="666666")
FONT_INSTRUCTION = Font(name="Calibri", size=10, italic=True, color="888888")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_WRAP = Alignment(wrap_text=True, vertical="center")

# Paleta por pilar
PILLAR_COLORS = {
    "P1": {"dark": "1F3864", "medium": "2F5496", "light": "D6E4F0"},
    "P2": {"dark": "2E75B6", "medium": "5B9BD5", "light": "DAEEF3"},
    "P3": {"dark": "375623", "medium": "548235", "light": "E2EFDA"},
    "P4": {"dark": "7030A0", "medium": "8B5FC7", "light": "E8D5F5"},
    "P5": {"dark": "843C0C", "medium": "BF6B30", "light": "FBE5D6"},
}

# Status colors
STATUS_FILLS = {
    "PLANEJADA": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "EM ANDAMENTO": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    "BLOQUEADA": PatternFill(start_color="FCD5D5", end_color="FCD5D5", fill_type="solid"),
    "CONCLUIDA": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "CANCELADA": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

PRIO_FILLS = {
    "P0": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "P1": PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid"),
    "P2": PatternFill(start_color="FFD43B", end_color="FFD43B", fill_type="solid"),
}
PRIO_FONTS = {
    "P0": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
    "P1": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
    "P2": Font(name="Calibri", size=11, bold=True, color="333333"),
}

# Column definitions (17 columns)
COLUMNS = [
    ("Codigo", 16),
    ("Pilar", 8),
    ("Subpilar", 10),
    ("Motor", 18),
    ("Titulo da Acao", 50),
    ("Tipo", 8),
    ("Prioridade", 12),
    ("OKR", 10),
    ("KR Principal", 14),
    ("Responsavel", 22),
    ("Patrocinador", 22),
    ("Prazo", 16),
    ("Custo (R$)", 18),
    ("Status", 16),
    ("Progresso %", 12),
    ("Evidencia (EVID-*)", 22),
    ("Observacoes", 40),
]

# ─────────────────────────────────────────────────────────
# DADOS CANONICOS PE2026
# ─────────────────────────────────────────────────────────

PILLARS = {
    "P1": "Governanca, Separacao Aero x TD e Padrao Auditavel",
    "P2": "Crescimento, Expansao e Diversificacao",
    "P3": "Excelencia Operacional e Escala com Margem",
    "P4": "Produto, Dados e IA como Vantagem Defensavel",
    "P5": "Pessoas, Lideranca e Capacidade Intelectual",
}

SUBPILLARS = {
    "P1.S1": "Separacao por unidades Aero x TD",
    "P1.S2": "Contabilidade gerencial e previsibilidade",
    "P1.S3": "Conformidade, contratos e padrao de auditoria",
    "P1.S4": "Gestao de riscos (regulatorio, operacional, reputacional)",
    "P2.S1": "Expansao TD (geografica e institucional)",
    "P2.S2": "Monetizacao da base contratual",
    "P2.S3": "Diversificacao Aero (fora TD)",
    "P2.S4": "Estrategia comercial e canais",
    "P3.S1": "Produtividade e capacidade operacional",
    "P3.S2": "Qualidade, padronizacao e reducao de retrabalho",
    "P3.S3": "Gestao de custos e escala com margem",
    "P3.S4": "Integracao Op <> CS <> Comercial <> Dados",
    "P4.S1": "Produto TD: evolucao, pacote de valor e retencao",
    "P4.S2": "Dados como evidencia (prova de valor)",
    "P4.S3": "IA aplicada e automacao",
    "P4.S4": "Arquitetura e governanca de dados",
    "P5.S1": "Lideranca e gestao (rituais, clareza, execucao)",
    "P5.S2": "Talentos e densidade intelectual",
    "P5.S3": "Cultura, engajamento e reconhecimento",
    "P5.S4": "People Analytics e governanca de RH",
}

MOTORS = {
    "M1": "M1 - Monetizacao",
    "M2": "M2 - Governanca",
    "M3": "M3 - Escala",
    "M4": "M4 - Produto/IA",
    "M5": "M5 - Pessoas",
}

# 22 INITs corporativas (DOC 08 v2)
CORPORATE_INITS = [
    # Motor M1 — Monetizacao
    {"code": "INIT-001", "pillar": "P2", "subpillar": "P2.S2", "motor": "M1", "title": "Implantar Sala de Situacao Q1 (Pareto Top-14)", "type": "MET", "priority": "P0", "okr": "OKR-P2", "kr": "P2.3, P2.5", "owner": "CS + Operacao", "sponsor": "Direcao Executiva", "deadline": "Fev-Mar/26", "cost": "R$ 5-10K", "evid": "EVID-2026-001", "areas": ["operacoes", "cs"]},
    {"code": "INIT-002", "pillar": "P2", "subpillar": "P2.S2", "motor": "M1", "title": "Criar Playbook de Ativacao de Demanda", "type": "MET", "priority": "P0", "okr": "OKR-P2", "kr": "P2.4, P2.5", "owner": "CS", "sponsor": "Direcao Executiva", "deadline": "Mar-Abr/26", "cost": "R$ 15-25K", "evid": "EVID-2026-002", "areas": ["cs"]},
    {"code": "INIT-003", "pillar": "P4", "subpillar": "P4.S3", "motor": "M1", "title": "Painel de monetizacao (saldo, vazao, idade, previsao, Pareto)", "type": "SIS", "priority": "P0", "okr": "OKR-P4", "kr": "P4.3", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Mar-Abr/26", "cost": "R$ 30-50K", "evid": "EVID-2026-003", "areas": ["pd"]},
    {"code": "INIT-004", "pillar": "P4", "subpillar": "P4.S2", "motor": "M1", "title": "Pacote mensal de prova de valor para Top-14", "type": "COM", "priority": "P0", "okr": "OKR-P4", "kr": "P4.1, P4.2", "owner": "Marketing + CS", "sponsor": "Direcao Executiva", "deadline": "Mar-Jun/26", "cost": "R$ 10-20K", "evid": "EVID-2026-004", "areas": ["marketing", "cs"]},
    {"code": "INIT-005", "pillar": "P3", "subpillar": "P3.S4", "motor": "M1", "title": "Integracao formal CS <> Operacao <> Financeiro", "type": "MET", "priority": "P0", "okr": "OKR-P3", "kr": "P3.5", "owner": "Operacao + CS", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 5-10K", "evid": "EVID-2026-005", "areas": ["operacoes", "cs"]},
    {"code": "INIT-017", "pillar": "P2", "subpillar": "P2.S1", "motor": "M1", "title": "Estrategia de feiras e eventos (envelope MKT)", "type": "COM", "priority": "P1", "okr": "OKR-P2", "kr": "P2.x", "owner": "Marketing", "sponsor": "Direcao Executiva", "deadline": "Mar-Dez/26", "cost": "R$ 500K", "evid": "EVID-2026-017", "areas": ["marketing"]},
    {"code": "INIT-018", "pillar": "P2", "subpillar": "P2.S1", "motor": "M1", "title": "Campanha 10 anos Techdengue", "type": "COM", "priority": "P1", "okr": "OKR-P2", "kr": "P2.x", "owner": "Marketing", "sponsor": "Direcao Executiva", "deadline": "Q2/26", "cost": "R$ 80-100K", "evid": "EVID-2026-018", "areas": ["marketing"]},
    {"code": "INIT-021", "pillar": "P2", "subpillar": "P2.S4", "motor": "M1", "title": "Estruturacao area Comercial (processos, CRM, pipeline)", "type": "ORG", "priority": "P1", "okr": "OKR-P2", "kr": "P2.x", "owner": "Lideranca Comercial", "sponsor": "Direcao Executiva", "deadline": "Mai-Jul/26", "cost": "R$ 40-70K", "evid": "EVID-2026-021", "areas": ["comercial"]},
    # Motor M2 — Governanca
    {"code": "INIT-006", "pillar": "P1", "subpillar": "P1.S3", "motor": "M2", "title": "Registro de decisoes (DEC-*) e riscos (RSK-*) ativo", "type": "MET", "priority": "P0", "okr": "OKR-P1", "kr": "P1.2, P1.3", "owner": "Direcao + Consultora", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 5-10K", "evid": "EVID-2026-006", "areas": ["financeiro"]},
    {"code": "INIT-007", "pillar": "P1", "subpillar": "P1.S2", "motor": "M2", "title": "Apuracao gerencial por unidade (Aero x TD)", "type": "MET", "priority": "P0", "okr": "OKR-P1", "kr": "P1.1", "owner": "Financeiro", "sponsor": "Direcao Executiva", "deadline": "Mar-Dez/26", "cost": "R$ 10-20K", "evid": "EVID-2026-007", "areas": ["financeiro"]},
    {"code": "INIT-019", "pillar": "P1", "subpillar": "P1.S1", "motor": "M2", "title": "Modelo separacao Aero x TD por fases (v1.0)", "type": "MET", "priority": "P1", "okr": "OKR-P1", "kr": "P1.5", "owner": "Direcao Executiva", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 20-40K", "evid": "EVID-2026-019", "areas": ["financeiro"]},
    # Motor M3 — Escala
    {"code": "INIT-008", "pillar": "P3", "subpillar": "P3.S1", "motor": "M3", "title": "Planejamento semanal de capacidade (Q1)", "type": "MET", "priority": "P0", "okr": "OKR-P3", "kr": "P3.2", "owner": "Operacao", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 3-5K", "evid": "EVID-2026-008", "areas": ["operacoes"]},
    {"code": "INIT-009", "pillar": "P3", "subpillar": "P3.S2", "motor": "M3", "title": "Padrao minimo de qualidade e reducao de retrabalho", "type": "ENT", "priority": "P0", "okr": "OKR-P3", "kr": "P3.3, P3.4", "owner": "Operacao", "sponsor": "Direcao Executiva", "deadline": "Mar-Jun/26", "cost": "R$ 15-25K", "evid": "EVID-2026-009", "areas": ["operacoes"]},
    # Motor M4 — Produto/IA
    {"code": "INIT-010", "pillar": "P4", "subpillar": "P4.S2", "motor": "M4", "title": "Padronizar relatorio executivo de evidencia (v1.0)", "type": "SIS", "priority": "P0", "okr": "OKR-P4", "kr": "P4.2", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 15-25K", "evid": "EVID-2026-010", "areas": ["pd"]},
    {"code": "INIT-011", "pillar": "P4", "subpillar": "P4.S3", "motor": "M4", "title": "1a automacao/IA com ganho operacional mensuravel", "type": "SIS", "priority": "P0", "okr": "OKR-P4", "kr": "P4.5", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Abr-Set/26", "cost": "R$ 40-70K", "evid": "EVID-2026-011", "areas": ["pd"]},
    {"code": "INIT-014", "pillar": "P4", "subpillar": "P4.S3", "motor": "M4", "title": "Centro de Tecnologia e IA", "type": "SIS", "priority": "P1", "okr": "OKR-P4", "kr": "P4.5", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Abr-Dez/26", "cost": "R$ 230-370K", "evid": "EVID-2026-014", "areas": ["pd"]},
    {"code": "INIT-020", "pillar": "P4", "subpillar": "P4.S1", "motor": "M4", "title": "Baseline de uso do produto + melhoria >= 15%", "type": "SIS", "priority": "P1", "okr": "OKR-P4", "kr": "P4.4", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Abr-Dez/26", "cost": "R$ 30-50K", "evid": "EVID-2026-020", "areas": ["pd"]},
    # Motor M5 — Pessoas
    {"code": "INIT-012", "pillar": "P5", "subpillar": "P5.S2", "motor": "M5", "title": "Mapa de posicoes-chave 2026 + plano preenchimento", "type": "ORG", "priority": "P0", "okr": "OKR-P5", "kr": "P5.4", "owner": "RH", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 5-10K", "evid": "EVID-2026-012", "areas": ["rh"]},
    {"code": "INIT-013", "pillar": "P5", "subpillar": "P5.S3", "motor": "M5", "title": "Onboarding estruturado (checklist + 45/90 dias)", "type": "MET", "priority": "P0", "okr": "OKR-P5", "kr": "P5.5", "owner": "RH", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 10-15K", "evid": "EVID-2026-013", "areas": ["rh"]},
    {"code": "INIT-015", "pillar": "P5", "subpillar": "P5.S2", "motor": "M5", "title": "Contratacao COO / Diretor de Operacoes", "type": "ORG", "priority": "P1", "okr": "OKR-P5", "kr": "P5.4", "owner": "RH + Direcao", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 216-264K/ano", "evid": "EVID-2026-015", "areas": ["rh"]},
    {"code": "INIT-016", "pillar": "P5", "subpillar": "P5.S2", "motor": "M5", "title": "Contratacao lideranca Comercial", "type": "ORG", "priority": "P1", "okr": "OKR-P5", "kr": "P5.4", "owner": "RH + Direcao", "sponsor": "Direcao Executiva", "deadline": "Mai/26", "cost": "R$ 144-192K/ano", "evid": "EVID-2026-016", "areas": ["rh"]},
    {"code": "INIT-022", "pillar": "P5", "subpillar": "P5.S1", "motor": "M5", "title": "Rituais minimos de lideranca (aderencia >= 85%)", "type": "MET", "priority": "P1", "okr": "OKR-P5", "kr": "P5.3", "owner": "RH + Direcao", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 10-20K", "evid": "EVID-2026-022", "areas": ["rh"]},
]

# INITs setoriais por area
AREA_INITS = {
    "rh": [
        {"code": "INIT-RH-301", "pillar": "P5", "subpillar": "P5.S3", "motor": "M5", "title": "Pesquisa de engajamento e clima 2026 (baseline)", "type": "MET", "priority": "P0", "okr": "OKR-P5", "kr": "P5.2", "owner": "RH", "sponsor": "Direcao Executiva", "deadline": "Mar-Abr/26", "cost": "R$ 5-10K", "evid": "EVID-RH-301"},
        {"code": "INIT-RH-302", "pillar": "P5", "subpillar": "P5.S4", "motor": "M5", "title": "Implantacao de indicadores de turnover por area", "type": "SIS", "priority": "P0", "okr": "OKR-P5", "kr": "P5.1", "owner": "RH", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 3-5K", "evid": "EVID-RH-302"},
        {"code": "INIT-RH-303", "pillar": "P5", "subpillar": "P5.S2", "motor": "M5", "title": "Processo seletivo estruturado para posicoes P0/P1", "type": "MET", "priority": "P0", "okr": "OKR-P5", "kr": "P5.4", "owner": "RH", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 5-8K", "evid": "EVID-RH-303"},
        {"code": "INIT-RH-304", "pillar": "P5", "subpillar": "P5.S3", "motor": "M5", "title": "Programa de reconhecimento e retencao", "type": "ENT", "priority": "P1", "okr": "OKR-P5", "kr": "P5.2", "owner": "RH", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 15-25K", "evid": "EVID-RH-304"},
        {"code": "INIT-RH-305", "pillar": "P5", "subpillar": "P5.S4", "motor": "M5", "title": "People Analytics: painel de pessoas (turnover, NPS, treinamento)", "type": "SIS", "priority": "P1", "okr": "OKR-P5", "kr": "P5.1", "owner": "RH", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 10-20K", "evid": "EVID-RH-305"},
        {"code": "INIT-RH-306", "pillar": "P5", "subpillar": "P5.S2", "motor": "M5", "title": "Programa de desenvolvimento de liderancas internas", "type": "MET", "priority": "P1", "okr": "OKR-P5", "kr": "P5.3", "owner": "RH", "sponsor": "Direcao Executiva", "deadline": "Abr-Set/26", "cost": "R$ 20-30K", "evid": "EVID-RH-306"},
    ],
    "marketing": [
        {"code": "INIT-MKT-101", "pillar": "P2", "subpillar": "P2.S1", "motor": "M1", "title": "Plano de comunicacao institucional 2026", "type": "MET", "priority": "P0", "okr": "OKR-P2", "kr": "P2.5", "owner": "Marketing", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 5-10K", "evid": "EVID-MKT-101"},
        {"code": "INIT-MKT-102", "pillar": "P2", "subpillar": "P2.S1", "motor": "M1", "title": "Calendario de feiras e eventos (envelope R$ 500K)", "type": "COM", "priority": "P0", "okr": "OKR-P2", "kr": "P2.x", "owner": "Marketing", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 500K (parte)", "evid": "EVID-MKT-102"},
        {"code": "INIT-MKT-103", "pillar": "P5", "subpillar": "P5.S2", "motor": "M5", "title": "Contratacao Marketing #1 e #2", "type": "ORG", "priority": "P0", "okr": "OKR-P5", "kr": "P5.4", "owner": "RH + Marketing", "sponsor": "Direcao Executiva", "deadline": "Mar-Abr/26", "cost": "R$ 95-152K/ano", "evid": "EVID-MKT-103"},
        {"code": "INIT-MKT-104", "pillar": "P4", "subpillar": "P4.S2", "motor": "M1", "title": "Pacote mensal de prova de valor (co-exec P&D/CS)", "type": "COM", "priority": "P0", "okr": "OKR-P4", "kr": "P4.1", "owner": "Marketing + CS", "sponsor": "Direcao Executiva", "deadline": "Mar-Jun/26", "cost": "R$ 8-15K", "evid": "EVID-MKT-104"},
        {"code": "INIT-MKT-105", "pillar": "P2", "subpillar": "P2.S1", "motor": "M1", "title": "Campanha 10 anos Techdengue (execucao)", "type": "COM", "priority": "P1", "okr": "OKR-P2", "kr": "P2.x", "owner": "Marketing", "sponsor": "Direcao Executiva", "deadline": "Q2/26", "cost": "R$ 80-100K", "evid": "EVID-MKT-105"},
        {"code": "INIT-MKT-106", "pillar": "P2", "subpillar": "P2.S4", "motor": "M1", "title": "Material de posicionamento institucional atualizado", "type": "COM", "priority": "P1", "okr": "OKR-P2", "kr": "P2.5", "owner": "Marketing", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 10-15K", "evid": "EVID-MKT-106"},
        {"code": "INIT-MKT-107", "pillar": "P4", "subpillar": "P4.S2", "motor": "M1", "title": "Estrategia de conteudo digital (blog, redes, relatorios)", "type": "MET", "priority": "P1", "okr": "OKR-P4", "kr": "P4.1", "owner": "Marketing", "sponsor": "Direcao Executiva", "deadline": "Abr-Dez/26", "cost": "R$ 10-20K", "evid": "EVID-MKT-107"},
    ],
    "pd": [
        {"code": "INIT-PD-251", "pillar": "P4", "subpillar": "P4.S4", "motor": "M4", "title": "Arquitetura de dados: modelo canonico e governanca", "type": "SIS", "priority": "P0", "okr": "OKR-P4", "kr": "P4.4", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Mar-Abr/26", "cost": "R$ 20-35K", "evid": "EVID-PD-251"},
        {"code": "INIT-PD-252", "pillar": "P4", "subpillar": "P4.S3", "motor": "M4", "title": "Dashboard interno metricas operacionais (v1.0)", "type": "SIS", "priority": "P0", "okr": "OKR-P4", "kr": "P4.3", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Mar-Abr/26", "cost": "R$ 15-25K", "evid": "EVID-PD-252"},
        {"code": "INIT-PD-253", "pillar": "P4", "subpillar": "P4.S1", "motor": "M4", "title": "Baseline de uso do produto Techdengue (metricas adocao)", "type": "SIS", "priority": "P0", "okr": "OKR-P4", "kr": "P4.4", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 10-15K", "evid": "EVID-PD-253"},
        {"code": "INIT-PD-254", "pillar": "P4", "subpillar": "P4.S1", "motor": "M4", "title": "Roadmap de produto 2026 com prioridades definidas", "type": "MET", "priority": "P1", "okr": "OKR-P4", "kr": "P4.4", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 5-10K", "evid": "EVID-PD-254"},
        {"code": "INIT-PD-255", "pillar": "P4", "subpillar": "P4.S2", "motor": "M4", "title": "Automacao de relatorio de prova de valor (v2.0)", "type": "SIS", "priority": "P1", "okr": "OKR-P4", "kr": "P4.2", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Mai-Set/26", "cost": "R$ 25-40K", "evid": "EVID-PD-255"},
        {"code": "INIT-PD-256", "pillar": "P4", "subpillar": "P4.S3", "motor": "M4", "title": "Estudo e PoC IA aplicada a operacao de campo", "type": "SIS", "priority": "P1", "okr": "OKR-P4", "kr": "P4.5", "owner": "Direcao + Consultoria", "sponsor": "Direcao Executiva", "deadline": "Jun-Set/26", "cost": "R$ 40-60K", "evid": "EVID-PD-256"},
    ],
    "operacoes": [
        {"code": "INIT-OP-151", "pillar": "P3", "subpillar": "P3.S1", "motor": "M3", "title": "Implantacao de agenda e planejamento semanal Q1", "type": "MET", "priority": "P0", "okr": "OKR-P3", "kr": "P3.2", "owner": "Operacao", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 3-5K", "evid": "EVID-OP-151"},
        {"code": "INIT-OP-152", "pillar": "P3", "subpillar": "P3.S2", "motor": "M3", "title": "Baseline de retrabalho por tipo de servico (mar/26)", "type": "SIS", "priority": "P0", "okr": "OKR-P3", "kr": "P3.3", "owner": "Operacao", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 3-5K", "evid": "EVID-OP-152"},
        {"code": "INIT-OP-153", "pillar": "P3", "subpillar": "P3.S2", "motor": "M3", "title": "Padrao minimo de SLA e qualidade por contratante", "type": "ENT", "priority": "P0", "okr": "OKR-P3", "kr": "P3.4", "owner": "Operacao", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 5-10K", "evid": "EVID-OP-153"},
        {"code": "INIT-OP-154", "pillar": "P2", "subpillar": "P2.S2", "motor": "M1", "title": "Estrutura de relatorio de execucao semanal (war room)", "type": "MET", "priority": "P0", "okr": "OKR-P2", "kr": "P2.3", "owner": "Operacao", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 2-3K", "evid": "EVID-OP-154"},
        {"code": "INIT-OP-155", "pillar": "P3", "subpillar": "P3.S4", "motor": "M3", "title": "Protocolo de interface com CS (pontos de passagem)", "type": "MET", "priority": "P0", "okr": "OKR-P3", "kr": "P3.5", "owner": "Operacao + CS", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 3-5K", "evid": "EVID-OP-155"},
        {"code": "INIT-OP-156", "pillar": "P3", "subpillar": "P3.S1", "motor": "M3", "title": "Planejamento de capacidade Q2/Q3 (pos COO)", "type": "MET", "priority": "P1", "okr": "OKR-P3", "kr": "P3.2", "owner": "Operacao (COO)", "sponsor": "Direcao Executiva", "deadline": "Mai-Jun/26", "cost": "R$ 5-10K", "evid": "EVID-OP-156"},
        {"code": "INIT-OP-157", "pillar": "P3", "subpillar": "P3.S2", "motor": "M3", "title": "Mapeamento e padronizacao de processos operacionais", "type": "MET", "priority": "P1", "okr": "OKR-P3", "kr": "P3.3", "owner": "Operacao", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 10-15K", "evid": "EVID-OP-157"},
        {"code": "INIT-OP-158", "pillar": "P3", "subpillar": "P3.S3", "motor": "M3", "title": "Indicadores de produtividade por equipe/regiao", "type": "SIS", "priority": "P1", "okr": "OKR-P3", "kr": "P3.1", "owner": "Operacao", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 8-12K", "evid": "EVID-OP-158"},
    ],
    "cs": [
        {"code": "INIT-CS-201", "pillar": "P2", "subpillar": "P2.S2", "motor": "M1", "title": "Mapeamento Pareto Top-14 com plano ativacao por cliente", "type": "MET", "priority": "P0", "okr": "OKR-P2", "kr": "P2.3", "owner": "CS", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 3-5K", "evid": "EVID-CS-201"},
        {"code": "INIT-CS-202", "pillar": "P2", "subpillar": "P2.S2", "motor": "M1", "title": "Implantacao de previsao 30/60/90 por contratante", "type": "SIS", "priority": "P0", "okr": "OKR-P2", "kr": "P2.3", "owner": "CS", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 5-8K", "evid": "EVID-CS-202"},
        {"code": "INIT-CS-203", "pillar": "P2", "subpillar": "P2.S2", "motor": "M1", "title": "Ritual semanal de war room com agenda padronizada", "type": "MET", "priority": "P0", "okr": "OKR-P2", "kr": "P2.5", "owner": "CS + Operacao", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 2-3K", "evid": "EVID-CS-203"},
        {"code": "INIT-CS-204", "pillar": "P2", "subpillar": "P2.S2", "motor": "M1", "title": "Playbook de ativacao de demanda (co-construcao)", "type": "MET", "priority": "P0", "okr": "OKR-P2", "kr": "P2.4", "owner": "CS", "sponsor": "Direcao Executiva", "deadline": "Mar-Abr/26", "cost": "R$ 10-15K", "evid": "EVID-CS-204"},
        {"code": "INIT-CS-205", "pillar": "P3", "subpillar": "P3.S4", "motor": "M3", "title": "SLA de atendimento e resposta ao contratante", "type": "ENT", "priority": "P0", "okr": "OKR-P3", "kr": "P3.4", "owner": "CS", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 3-5K", "evid": "EVID-CS-205"},
        {"code": "INIT-CS-206", "pillar": "P4", "subpillar": "P4.S2", "motor": "M1", "title": "Base de conhecimento de clientes Pareto (perfil+historico)", "type": "SIS", "priority": "P1", "okr": "OKR-P4", "kr": "P4.1", "owner": "CS", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 8-12K", "evid": "EVID-CS-206"},
        {"code": "INIT-CS-207", "pillar": "P4", "subpillar": "P4.S2", "motor": "M1", "title": "Protocolo de entrega de prova de valor mensalmente", "type": "MET", "priority": "P1", "okr": "OKR-P4", "kr": "P4.2", "owner": "CS + Marketing", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 3-5K", "evid": "EVID-CS-207"},
        {"code": "INIT-CS-208", "pillar": "P2", "subpillar": "P2.S1", "motor": "M1", "title": "Estruturacao de pipeline de renovacoes e expansao", "type": "MET", "priority": "P1", "okr": "OKR-P2", "kr": "P2.1", "owner": "CS", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 5-8K", "evid": "EVID-CS-208"},
    ],
    "comercial": [
        {"code": "INIT-COM-401", "pillar": "P2", "subpillar": "P2.S4", "motor": "M1", "title": "Definicao de processos comerciais e pipeline", "type": "MET", "priority": "P1", "okr": "OKR-P2", "kr": "P2.x", "owner": "Lideranca Comercial", "sponsor": "Direcao Executiva", "deadline": "Jun-Jul/26", "cost": "R$ 10-15K", "evid": "EVID-COM-401"},
        {"code": "INIT-COM-402", "pillar": "P2", "subpillar": "P2.S4", "motor": "M1", "title": "Implantacao de CRM (selecao + configuracao)", "type": "SIS", "priority": "P1", "okr": "OKR-P2", "kr": "P2.x", "owner": "Lideranca Comercial", "sponsor": "Direcao Executiva", "deadline": "Jun-Ago/26", "cost": "R$ 15-25K", "evid": "EVID-COM-402"},
        {"code": "INIT-COM-403", "pillar": "P2", "subpillar": "P2.S4", "motor": "M1", "title": "Playbook comercial v1.0 (proposta, negociacao, fechamento)", "type": "MET", "priority": "P1", "okr": "OKR-P2", "kr": "P2.4", "owner": "Lideranca Comercial", "sponsor": "Direcao Executiva", "deadline": "Jul-Ago/26", "cost": "R$ 8-12K", "evid": "EVID-COM-403"},
        {"code": "INIT-COM-404", "pillar": "P2", "subpillar": "P2.S3", "motor": "M1", "title": "Mapeamento de prospects Aero (diversificacao)", "type": "MET", "priority": "P1", "okr": "OKR-P2", "kr": "P2.x", "owner": "Lideranca Comercial", "sponsor": "Direcao Executiva", "deadline": "Jul-Set/26", "cost": "R$ 5-8K", "evid": "EVID-COM-404"},
        {"code": "INIT-COM-405", "pillar": "P3", "subpillar": "P3.S4", "motor": "M3", "title": "Integracao Comercial <> CS <> Marketing", "type": "MET", "priority": "P2", "okr": "OKR-P3", "kr": "P3.5", "owner": "Lideranca Comercial", "sponsor": "Direcao Executiva", "deadline": "Q3/26", "cost": "R$ 3-5K", "evid": "EVID-COM-405"},
    ],
    "financeiro": [
        {"code": "INIT-FIN-351", "pillar": "P1", "subpillar": "P1.S2", "motor": "M2", "title": "DRE gerencial por unidade (Aero x TD) - implantacao", "type": "ENT", "priority": "P0", "okr": "OKR-P1", "kr": "P1.1", "owner": "Financeiro", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 8-12K", "evid": "EVID-FIN-351"},
        {"code": "INIT-FIN-352", "pillar": "P1", "subpillar": "P1.S2", "motor": "M2", "title": "Centros de custo e alcadas formalizados", "type": "MET", "priority": "P0", "okr": "OKR-P1", "kr": "P1.2", "owner": "Financeiro", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 3-5K", "evid": "EVID-FIN-352"},
        {"code": "INIT-FIN-353", "pillar": "P1", "subpillar": "P1.S2", "motor": "M2", "title": "Projecao de caixa 30/60/90 semanal (rotina)", "type": "MET", "priority": "P0", "okr": "OKR-P1", "kr": "G-02", "owner": "Financeiro", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 2-3K", "evid": "EVID-FIN-353"},
        {"code": "INIT-FIN-354", "pillar": "P1", "subpillar": "P1.S2", "motor": "M2", "title": "Dashboard financeiro gerencial (margem, receita, caixa)", "type": "SIS", "priority": "P0", "okr": "OKR-P1", "kr": "P1.1", "owner": "Financeiro", "sponsor": "Direcao Executiva", "deadline": "Mar-Abr/26", "cost": "R$ 10-15K", "evid": "EVID-FIN-354"},
        {"code": "INIT-FIN-355", "pillar": "P1", "subpillar": "P1.S4", "motor": "M2", "title": "Registro de riscos financeiros e gatilhos (RSK-*)", "type": "MET", "priority": "P0", "okr": "OKR-P1", "kr": "P1.3", "owner": "Financeiro", "sponsor": "Direcao Executiva", "deadline": "Mar/26", "cost": "R$ 2-3K", "evid": "EVID-FIN-355"},
        {"code": "INIT-FIN-356", "pillar": "P1", "subpillar": "P1.S2", "motor": "M2", "title": "Relatorio mensal de desempenho financeiro (MBR)", "type": "MET", "priority": "P1", "okr": "OKR-P1", "kr": "P1.1", "owner": "Financeiro", "sponsor": "Direcao Executiva", "deadline": "Abr/26", "cost": "R$ 3-5K", "evid": "EVID-FIN-356"},
        {"code": "INIT-FIN-357", "pillar": "P1", "subpillar": "P1.S2", "motor": "M2", "title": "Orcamento trimestral rolling + revisao cenarios", "type": "MET", "priority": "P1", "okr": "OKR-P3", "kr": "P3.1", "owner": "Financeiro", "sponsor": "Direcao Executiva", "deadline": "Abr-Jun/26", "cost": "R$ 5-8K", "evid": "EVID-FIN-357"},
        {"code": "INIT-FIN-358", "pillar": "P1", "subpillar": "P1.S3", "motor": "M2", "title": "Compliance fiscal e documental para padrao auditavel", "type": "ENT", "priority": "P1", "okr": "OKR-P1", "kr": "P1.4", "owner": "Financeiro", "sponsor": "Direcao Executiva", "deadline": "Abr-Dez/26", "cost": "R$ 10-15K", "evid": "EVID-FIN-358"},
    ],
}

# Area metadata
AREAS = {
    "rh": {
        "name": "RH / Pessoas",
        "code": "11-A",
        "mandate": "Construir a infraestrutura de pessoas para sustentar crescimento com metodo, lideranca e densidade intelectual.",
        "okrs": "P5 (primario) - P1 (secundario)",
        "envelope": "R$ 630K",
        "filename": "PE2026_RH.xlsx",
    },
    "marketing": {
        "name": "Marketing",
        "code": "11-B",
        "mandate": "Gerar evidencia e posicionamento que ativam demanda, sustentam permanencia e preparam o Techdengue como marca transacionavel.",
        "okrs": "P2 (primario) - P4 (secundario)",
        "envelope": "R$ 1.280K",
        "filename": "PE2026_Marketing.xlsx",
    },
    "pd": {
        "name": "P&D / Produto / Dados",
        "code": "11-C",
        "mandate": "Transformar tecnologia, dados e produto em vantagem defensavel via Direcao Executiva + consultorias especializadas.",
        "okrs": "P4 (primario)",
        "envelope": "R$ 801K + R$ 300K (IA)",
        "filename": "PE2026_PD.xlsx",
    },
    "operacoes": {
        "name": "Operacao",
        "code": "11-D",
        "mandate": "Executar com qualidade, produtividade e previsibilidade, sustentando crescimento e monetizacao sem comprometer margem.",
        "okrs": "P2 (primario) - P3 (primario)",
        "envelope": "R$ 1.062K",
        "filename": "PE2026_Operacao.xlsx",
    },
    "cs": {
        "name": "CS / Relacionamento",
        "code": "11-E",
        "mandate": "Conectar contrato a demanda e execucao - a area que ativa e retem a base contratual do Techdengue.",
        "okrs": "P2 (primario) - P3 (secundario)",
        "envelope": "Parte do envelope Comercial",
        "filename": "PE2026_CS.xlsx",
    },
    "comercial": {
        "name": "Comercial (area em criacao)",
        "code": "11-F",
        "mandate": "Estruturar a area Comercial do zero - metodo, pipeline, CRM e posicionamento - a partir de mai/2026.",
        "okrs": "P2 (primario, condicional)",
        "envelope": "R$ 820K",
        "filename": "PE2026_Comercial.xlsx",
    },
    "financeiro": {
        "name": "Financeiro",
        "code": "11-G",
        "mandate": "Garantir margem, previsibilidade, controles e disciplina de gestao financeira - base para padrao auditavel e sell-ready.",
        "okrs": "P1 (primario) - P3 (secundario)",
        "envelope": "R$ 253K + parte Estrategico",
        "filename": "PE2026_Financeiro.xlsx",
    },
}

AREA_PRIMARY_PILLARS = {
    "rh": ("P5", "P1"),
    "marketing": ("P2", "P4"),
    "pd": ("P4",),
    "operacoes": ("P3", "P2"),
    "cs": ("P2", "P3"),
    "comercial": ("P2", "P3"),
    "financeiro": ("P1", "P3"),
}

AREA_INTERFACES = {
    "rh": "Direcao Executiva, Financeiro e liderancas de area",
    "marketing": "Comercial, P&D / Produto / Dados e CS / Relacionamento",
    "pd": "Direcao Executiva, Marketing e CS / Relacionamento",
    "operacoes": "CS / Relacionamento, Financeiro e P&D / Produto / Dados",
    "cs": "Operacao, Financeiro e Direcao Executiva",
    "comercial": "Marketing, CS / Relacionamento e Direcao Executiva",
    "financeiro": "Direcao Executiva, Operacao e RH / Pessoas",
}

ACTION_STEP_TEMPLATES = {
    "MET": [
        {
            "title": "Diagnostico e baseline",
            "deliverable": "Baseline validado, riscos mapeados e situacao atual documentada",
        },
        {
            "title": "Desenho do padrao",
            "deliverable": "RACI, rotina, criterios e artefatos-padrao publicados",
        },
        {
            "title": "Piloto controlado",
            "deliverable": "Piloto em frente prioritaria com aprendizados e ajustes registrados",
        },
        {
            "title": "Estabilizacao e rotina",
            "deliverable": "Adoção estabilizada e ritual de acompanhamento institucionalizado",
        },
    ],
    "SIS": [
        {
            "title": "Requisitos, fontes e metricas",
            "deliverable": "Escopo, fontes, metricas e criterios de sucesso fechados",
        },
        {
            "title": "Arquitetura e integracao",
            "deliverable": "Mapa tecnico, integracoes e regras de conciliacao publicados",
        },
        {
            "title": "Construcao e homologacao",
            "deliverable": "Solucao configurada, testada e pronta para validacao com usuarios",
        },
        {
            "title": "Go-live e monitoramento",
            "deliverable": "Entrega em uso, monitorada com rotina de correcoes e estabilidade",
        },
        {
            "title": "Auditoria e evolucao",
            "deliverable": "Licoes aprendidas, backlog de melhoria e governanca de evolucao",
        },
    ],
    "COM": [
        {
            "title": "Tese, ICP e posicionamento",
            "deliverable": "Tese comercial, ICP e narrativa de valor validados com a direcao",
        },
        {
            "title": "Materiais e oferta",
            "deliverable": "Kit de abordagem, proposta e materiais de apoio padronizados",
        },
        {
            "title": "Ativacao de campo",
            "deliverable": "Acao em campo com pipeline, agenda e follow-up rastreado",
        },
        {
            "title": "Conversao e aprendizado",
            "deliverable": "Resultados consolidados, ajustes de tese e proximos passos definidos",
        },
    ],
    "ENT": [
        {
            "title": "Mapeamento e padrao minimo",
            "deliverable": "Frentes criticas, requisitos e padrao minimo documentados",
        },
        {
            "title": "Implantacao nas frentes criticas",
            "deliverable": "Padrao aplicado nos pontos de maior impacto operacional",
        },
        {
            "title": "Treinamento e adocao",
            "deliverable": "Equipe treinada e aderencia acompanhada em rotina",
        },
        {
            "title": "Auditoria e ajuste fino",
            "deliverable": "Auditoria, correcoes e estabilizacao do novo padrao",
        },
    ],
    "ORG": [
        {
            "title": "Desenho organizacional e alçadas",
            "deliverable": "Papeis, alçadas, fronteiras e responsabilidades definidos",
        },
        {
            "title": "Selecao, configuracao e contratacao",
            "deliverable": "Estrutura operacional configurada ou contratacao encaminhada",
        },
        {
            "title": "Onboarding e transicao",
            "deliverable": "Onboarding estruturado e transicao acompanhada por ritos",
        },
        {
            "title": "Governanca e desempenho",
            "deliverable": "Rotina de gestao, indicadores e performance institucionalizados",
        },
    ],
}


# ─────────────────────────────────────────────────────────
# FUNCOES DE CONSTRUCAO
# ─────────────────────────────────────────────────────────

def setup_config_sheet(wb):
    """Cria aba Config (oculta) com listas de validacao."""
    ws = wb.create_sheet("Config")

    lists_data = {
        "A": ("Status", ["PLANEJADA", "EM ANDAMENTO", "BLOQUEADA", "CONCLUIDA", "CANCELADA"]),
        "B": ("Tipo", ["ENT", "MET", "SIS", "ORG", "COM"]),
        "C": ("Prioridade", ["P0", "P1", "P2"]),
        "D": ("Motor", [MOTORS[k] for k in sorted(MOTORS.keys())]),
        "E": ("Pilar", list(PILLARS.keys())),
        "F": ("Subpilar", sorted(SUBPILLARS.keys())),
    }

    for col, (header, values) in lists_data.items():
        ws[f"{col}1"] = header
        ws[f"{col}1"].font = Font(bold=True)
        for i, v in enumerate(values, start=2):
            ws[f"{col}{i}"] = v

    ws.sheet_state = "hidden"
    return ws


def add_conditional_formatting(ws, data_start_row, data_end_row):
    """Adiciona formatacao condicional por Status e Prioridade."""
    status_col = "N"  # col 14 = Status
    prio_col = "G"    # col 7 = Prioridade

    status_range = f"{status_col}{data_start_row}:{status_col}{data_end_row}"
    prio_range = f"{prio_col}{data_start_row}:{prio_col}{data_end_row}"

    for status, fill in STATUS_FILLS.items():
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill)
        )

    for prio, fill in PRIO_FILLS.items():
        ws.conditional_formatting.add(
            prio_range,
            CellIsRule(operator="equal", formula=[f'"{prio}"'], fill=fill,
                       font=PRIO_FONTS[prio])
        )


def add_data_validations(ws, data_start_row, data_end_row):
    """Adiciona listas suspensas (data validation)."""
    validations = {
        "B": "Config!$E$2:$E$6",    # Pilar
        "C": "Config!$F$2:$F$21",   # Subpilar
        "D": "Config!$D$2:$D$6",    # Motor
        "F": "Config!$B$2:$B$6",    # Tipo
        "G": "Config!$C$2:$C$4",    # Prioridade
        "N": "Config!$A$2:$A$6",    # Status
    }
    for col_letter, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Selecione um valor da lista"
        dv.prompt = "Escolha uma opcao"
        dv.sqref = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
        ws.add_data_validation(dv)

    # Progresso 0-100
    dv_prog = DataValidation(type="whole", operator="between", formula1="0", formula2="100")
    dv_prog.error = "Valor entre 0 e 100"
    dv_prog.sqref = f"O{data_start_row}:O{data_end_row}"
    ws.add_data_validation(dv_prog)


def write_header_row(ws, row):
    """Escreve cabecalho das 17 colunas."""
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = FONT_HEADER
        cell.fill = header_fill
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_pillar_group_header(ws, row, pillar_code, is_subpillar=False):
    """Escreve cabecalho de grupo (pilar ou subpilar)."""
    colors = PILLAR_COLORS.get(pillar_code[:2], PILLAR_COLORS["P1"])

    if is_subpillar:
        label = f"  {pillar_code} - {SUBPILLARS.get(pillar_code, pillar_code)}"
        fill_color = colors["medium"]
        font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    else:
        label = f"{pillar_code} - {PILLARS.get(pillar_code, pillar_code)}"
        fill_color = colors["dark"]
        font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")

    for c in range(2, len(COLUMNS) + 1):
        ws.cell(row=row, column=c).fill = fill


def write_init_row(ws, row, init_data, level="parent"):
    """Escreve uma linha de INIT ou subacao."""
    values = [
        init_data.get("code", ""),
        init_data.get("pillar", ""),
        init_data.get("subpillar", ""),
        MOTORS.get(init_data.get("motor", ""), init_data.get("motor", "")),
        init_data.get("title", ""),
        init_data.get("type", ""),
        init_data.get("priority", ""),
        init_data.get("okr", ""),
        init_data.get("kr", ""),
        init_data.get("owner", ""),
        init_data.get("sponsor", ""),
        init_data.get("deadline", ""),
        init_data.get("cost", ""),
        "PLANEJADA",
        0,
        init_data.get("evid", ""),
        init_data.get("notes", ""),
    ]

    pillar_code = init_data.get("pillar", "P1")[:2]
    palette = PILLAR_COLORS.get(pillar_code, PILLAR_COLORS["P1"])
    fill_color = palette["medium"] if level == "parent" else palette["light"]
    text_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF") if level == "parent" else FONT_BODY
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = text_font
        cell.border = THIN_BORDER
        cell.fill = row_fill
        if level == "child" and col_idx == 5:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        elif col_idx in (2, 3, 6, 7, 8, 14, 15):
            cell.alignment = ALIGN_CENTER
        else:
            cell.alignment = ALIGN_WRAP

    ws.row_dimensions[row].height = 26 if level == "parent" else 22


def build_detailed_steps(init_data):
    """Desdobra uma INIT em subacoes mais concretas por tipo."""
    templates = ACTION_STEP_TEMPLATES.get(init_data.get("type"), ACTION_STEP_TEMPLATES["MET"])
    detailed_steps = []

    for idx, template in enumerate(templates, start=1):
        detailed_steps.append({
            "code": f"{init_data['code']}-{idx:02d}",
            "pillar": init_data.get("pillar", ""),
            "subpillar": init_data.get("subpillar", ""),
            "motor": init_data.get("motor", ""),
            "title": template["title"],
            "type": init_data.get("type", ""),
            "priority": init_data.get("priority", ""),
            "okr": init_data.get("okr", ""),
            "kr": init_data.get("kr", ""),
            "owner": init_data.get("owner", ""),
            "sponsor": init_data.get("sponsor", ""),
            "deadline": init_data.get("deadline", ""),
            "cost": "",
            "status": "PLANEJADA",
            "progress": 0,
            "evid": f"{init_data.get('evid', 'EVID')}-{idx:02d}",
            "notes": f"Subacao {idx}/{len(templates)} | {template['deliverable']} | Pai: {init_data['code']}",
        })

    return detailed_steps


def get_inits_for_area(area_slug):
    """Retorna INITs corporativas + setoriais para uma area, agrupadas por pilar/subpilar."""
    corp_inits = [i for i in CORPORATE_INITS if area_slug in i.get("areas", [])]
    area_inits = AREA_INITS.get(area_slug, [])
    all_inits = corp_inits + area_inits

    grouped = {}
    for init in all_inits:
        p = init["pillar"]
        sp = init["subpillar"]
        if p not in grouped:
            grouped[p] = {}
        if sp not in grouped[p]:
            grouped[p][sp] = []
        grouped[p][sp].append(init)

    for p in grouped:
        for sp in grouped[p]:
            grouped[p][sp].sort(key=lambda x: (0 if x["priority"] == "P0" else 1 if x["priority"] == "P1" else 2, x["code"]))

    return grouped


def build_action_plan_sheet(wb, area_slug, area_meta):
    """Constroi a aba Plano de Acao com agrupamento por pilar/subpilar e subacoes."""
    ws = wb.create_sheet("Plano de Acao")
    ws.sheet_properties.tabColor = PILLAR_COLORS.get("P3", {}).get("dark", "548235")
    ws.sheet_properties.outlinePr.summaryBelow = True
    try:
        ws.sheet_view.showOutlineSymbols = True
    except Exception:
        pass

    # Cabecalho fixo
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    cell = ws.cell(row=row, column=1,
                   value=f"PE2026 - Execucao Detalhada: {area_meta['name']} ({area_meta['code']})")
    cell.font = FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    ws.cell(row=row, column=1, value=f"Mandato: {area_meta['mandate']}").font = FONT_BODY

    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    ws.cell(row=row, column=1,
            value=f"OKRs: {area_meta['okrs']}  |  Envelope: {area_meta['envelope']}  |  Versao 1.0 - Mar/2026").font = FONT_BODY

    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    ws.cell(row=row, column=1,
            value="Legenda: linha-mãe = INIT | linhas aninhadas = subacoes concretas | P0 = Critico 90 dias | P1 = Estrategico do ano | P2 = Melhoria").font = FONT_INSTRUCTION

    row = 5  # separador
    row = 6
    write_header_row(ws, row)

    data_start_row = 7
    current_row = 7

    grouped = get_inits_for_area(area_slug)

    for pillar_code in sorted(grouped.keys()):
        write_pillar_group_header(ws, current_row, pillar_code, is_subpillar=False)
        current_row += 1

        for subpillar_code in sorted(grouped[pillar_code].keys()):
            write_pillar_group_header(ws, current_row, subpillar_code, is_subpillar=True)
            current_row += 1

            for init in grouped[pillar_code][subpillar_code]:
                parent_init = dict(init)
                parent_init["notes"] = "Linha-mãe da iniciativa; subacoes detalhadas logo abaixo."
                write_init_row(ws, current_row, parent_init, level="parent")
                current_row += 1

                for step in build_detailed_steps(init):
                    write_init_row(ws, current_row, step, level="child")
                    ws.row_dimensions[current_row].outlineLevel = 1
                    current_row += 1

        current_row += 1  # espaco entre pilares

    data_end_row = current_row + 40  # margem para novas acoes

    # Rodape
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(COLUMNS))
    ws.cell(row=current_row, column=1,
            value="Fonte: DOC 08 v2 + DOC 06 v2 + DOC 04 v2 | PE2026 | Confidencial | Linha-mãe = INIT | Subacoes = desdobramento operacional").font = FONT_INSTRUCTION

    add_data_validations(ws, data_start_row, data_end_row)
    add_conditional_formatting(ws, data_start_row, data_end_row)

    # Freeze top rows
    ws.freeze_panes = "A7"

    return data_start_row, current_row - 1


def build_dashboard_sheet(wb, area_slug, area_meta):
    """Constroi a aba Dashboard com metricas automaticas."""
    ws = wb.active
    ws.title = "Dashboard"

    ws.sheet_properties.tabColor = PILLAR_COLORS.get("P2", {}).get("dark", "2E75B6")

    # Titulo
    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1,
                   value=f"PE2026 - Plano de Acao: {area_meta['name']}")
    cell.font = Font(name="Calibri", size=18, bold=True, color="1F3864")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1,
            value="Confidencial - Uso interno da Direcao | Mar/2026").font = FONT_INSTRUCTION

    ws.merge_cells("A3:H3")
    ws.cell(row=3, column=1,
            value=f"Mandato: {area_meta['mandate']}").font = FONT_BODY
    ws.row_dimensions[3].height = 30

    ws.merge_cells("A4:H4")
    ws.cell(row=4, column=1,
            value=f"OKRs: {area_meta['okrs']}  |  Envelope: {area_meta['envelope']}").font = FONT_SUBTITLE

    # KPI cards row
    row = 6
    card_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    card_border = Border(
        left=Side(style="medium", color="D9D9D9"),
        right=Side(style="medium", color="D9D9D9"),
        top=Side(style="medium", color="D9D9D9"),
        bottom=Side(style="medium", color="D9D9D9"),
    )

    cards = [
        ("A", "B", "TOTAL ACOES", '=COUNTIF(\'Plano de Acao\'!A7:A600,"INIT*")'),
        ("C", "D", "EM ANDAMENTO", '=COUNTIF(\'Plano de Acao\'!N:N,"EM ANDAMENTO")'),
        ("E", "F", "BLOQUEADAS", '=COUNTIF(\'Plano de Acao\'!N:N,"BLOQUEADA")'),
        ("G", "H", "CONCLUIDAS", '=COUNTIF(\'Plano de Acao\'!N:N,"CONCLUIDA")'),
    ]

    for col_start, col_end, label, formula in cards:
        ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
        c = ws.cell(row=row, column=ws[f"{col_start}1"].column, value=label)
        c.font = FONT_DASH_LABEL
        c.fill = card_fill
        c.alignment = ALIGN_CENTER
        c.border = card_border
        ws.cell(row=row, column=ws[f"{col_end}1"].column).fill = card_fill
        ws.cell(row=row, column=ws[f"{col_end}1"].column).border = card_border

        ws.merge_cells(f"{col_start}{row+1}:{col_end}{row+1}")
        c2 = ws.cell(row=row+1, column=ws[f"{col_start}1"].column, value=formula)
        c2.font = FONT_DASH_BIG
        c2.fill = card_fill
        c2.alignment = ALIGN_CENTER
        c2.border = card_border
        ws.cell(row=row+1, column=ws[f"{col_end}1"].column).fill = card_fill
        ws.cell(row=row+1, column=ws[f"{col_end}1"].column).border = card_border

    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 45

    # P0 abertas
    row = 9
    ws.merge_cells(f"A{row}:B{row}")
    ws.cell(row=row, column=1, value="ACOES P0 ABERTAS:").font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    ws.cell(row=row, column=3,
            value='=COUNTIFS(\'Plano de Acao\'!G:G,"P0",\'Plano de Acao\'!N:N,"<>CONCLUIDA",\'Plano de Acao\'!N:N,"<>CANCELADA")').font = Font(name="Calibri", size=14, bold=True, color="FF0000")

    # Progresso por pilar
    row = 11
    ws.cell(row=row, column=1, value="PROGRESSO POR PILAR").font = FONT_SUBTITLE

    for i, (pk, pname) in enumerate(sorted(PILLARS.items())):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=f"{pk} - {pname[:40]}").font = FONT_BODY
        ws.cell(row=r, column=5,
                value=f'=IFERROR(COUNTIFS(\'Plano de Acao\'!B:B,"{pk}",\'Plano de Acao\'!N:N,"CONCLUIDA")/COUNTIF(\'Plano de Acao\'!B:B,"{pk}")*100,0)').font = FONT_BODY
        ws.cell(row=r, column=5).number_format = '0"%"'
        ws.cell(row=r, column=6,
                value=f'=COUNTIF(\'Plano de Acao\'!B:B,"{pk}")').font = Font(name="Calibri", size=10, color="888888")

    # Legenda
    row = 18
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1,
            value="P0 = Critico 90 dias (vermelho) | P1 = Estrategico do ano (laranja) | P2 = Melhoria (amarelo)").font = FONT_INSTRUCTION

    # Column widths
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 16


def build_gerencial_sheet(wb, area_slug, area_meta):
    """Constroi a aba Gerencial com mandato, fronteiras e governanca."""
    grouped = get_inits_for_area(area_slug)
    all_inits = [init for pillar in grouped.values() for subpillar in pillar.values() for init in subpillar]
    priority_counts = {"P0": 0, "P1": 0, "P2": 0}
    for init in all_inits:
        priority_counts[init["priority"]] = priority_counts.get(init["priority"], 0) + 1

    ws = wb.create_sheet("Gerencial")
    ws.sheet_properties.tabColor = PILLAR_COLORS.get("P5", {}).get("dark", "843C0C")

    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1, value=f"PE2026 - Caderno Gerencial: {area_meta['name']} ({area_meta['code']})")
    cell.font = FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value=f"Mandato, fronteiras, interfaces e governanca | Versao 2.0 | {area_meta['envelope']}").font = FONT_INSTRUCTION

    ws.merge_cells("A4:H4")
    ws.cell(row=4, column=1, value=area_meta["mandate"]).font = FONT_SUBTITLE
    ws.row_dimensions[4].height = 30

    card_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    card_border = Border(
        left=Side(style="medium", color="D9D9D9"),
        right=Side(style="medium", color="D9D9D9"),
        top=Side(style="medium", color="D9D9D9"),
        bottom=Side(style="medium", color="D9D9D9"),
    )
    cards = [
        ("A", "B", "PILARES PRIMARIOS", " / ".join(AREA_PRIMARY_PILLARS.get(area_slug, ()))),
        ("C", "D", "INTERFACES CRITICAS", AREA_INTERFACES.get(area_slug, "")),
        ("E", "F", "ENVELOPE", area_meta["envelope"]),
        ("G", "H", "CADENCIA", "WBR | MBR | QBR"),
    ]

    row = 6
    for col_start, col_end, label, value in cards:
        ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
        c = ws.cell(row=row, column=ws[f"{col_start}1"].column, value=label)
        c.font = FONT_DASH_LABEL
        c.fill = card_fill
        c.alignment = ALIGN_CENTER
        c.border = card_border
        ws.cell(row=row, column=ws[f"{col_end}1"].column).fill = card_fill
        ws.cell(row=row, column=ws[f"{col_end}1"].column).border = card_border

        ws.merge_cells(f"{col_start}{row+1}:{col_end}{row+1}")
        c2 = ws.cell(row=row + 1, column=ws[f"{col_start}1"].column, value=value)
        c2.font = Font(name="Calibri", size=12, bold=True, color="1F3864")
        c2.fill = card_fill
        c2.alignment = ALIGN_CENTER
        c2.border = card_border
        ws.cell(row=row + 1, column=ws[f"{col_end}1"].column).fill = card_fill
        ws.cell(row=row + 1, column=ws[f"{col_end}1"].column).border = card_border

    row = 9
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="GOVERNANCA E RITOS").font = FONT_SUBTITLE

    row += 1
    governance_headers = ["Rito", "Cadencia", "Objetivo", "Registro"]
    governance_rows = [
        ["WBR", "Semanal", "Remover bloqueios, revisar P0 e decidir ajustes rapidos", "Ata resumida + DEC-* se necessario"],
        ["MBR", "Mensal", "Rever carteira, riscos, custos e resultados", "Relatorio mensal + log de decisoes"],
        ["QBR", "Trimestral", "Rever tese, capacidade e rebalancear carteira", "Relatorio trimestral + rebalanceamento"],
    ]
    for col_idx, header in enumerate(governance_headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    for col in "ABCD":
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 28

    for idx, gov in enumerate(governance_rows, start=1):
        r = row + idx
        for col_idx, val in enumerate(gov, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx < 3 else ALIGN_WRAP

    row += len(governance_rows) + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="REGRAS NAO NEGOCIAVEIS").font = FONT_SUBTITLE

    row += 1
    rule_headers = ["Regra", "Aplicacao"]
    rule_rows = [
        ["Custo por INIT", "Toda iniciativa deve ter ordem de grandeza em R$ conforme DOC 08/09"],
        ["Evidencia", "Nenhuma INIT encerra sem EVID verificavel e rastreavel"],
        ["WIP", "Maximo de 5 iniciativas ativas simultaneamente por area"],
        ["Criterio de corte", "P0 e P1 sao priorizadas; P2 nao compete com o motor principal"],
    ]
    for col_idx, header in enumerate(rule_headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 48
    for idx, rule in enumerate(rule_rows, start=1):
        r = row + idx
        for col_idx, val in enumerate(rule, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_WRAP

    row += len(rule_rows) + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="RESUMO DA CARTEIRA").font = FONT_SUBTITLE

    row += 1
    summary_headers = ["Prioridade", "Quantidade", "Leitura"]
    summary_rows = [
        ["P0", priority_counts.get("P0", 0), "Criticos de 90 dias e motor do ano"],
        ["P1", priority_counts.get("P1", 0), "Estruturantes do ano e consolidacao"],
        ["P2", priority_counts.get("P2", 0), "Melhorias e reforcos de suporte"],
    ]
    for col_idx, header in enumerate(summary_headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 44
    for idx, item in enumerate(summary_rows, start=1):
        r = row + idx
        for col_idx, val in enumerate(item, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx < 3 else ALIGN_WRAP

    ws.freeze_panes = "A10"


def build_direction_sheet(wb, area_slug, area_meta):
    """Constroi a aba Direcao com a carteira priorizada da area."""
    grouped = get_inits_for_area(area_slug)
    all_inits = [init for pillar in grouped.values() for subpillar in pillar.values() for init in subpillar]
    all_inits = sorted(all_inits, key=lambda x: (0 if x["priority"] == "P0" else 1 if x["priority"] == "P1" else 2, x["code"]))

    ws = wb.create_sheet("Direcao")
    ws.sheet_properties.tabColor = PILLAR_COLORS.get("P2", {}).get("dark", "2E75B6")

    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1, value=f"PE2026 - Caderno de Direcao: {area_meta['name']} ({area_meta['code']})")
    cell.font = FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value=f"OKRs orientadores: {area_meta['okrs']} | Carteira priorizada por pilar/subpilar | Versao 2.0").font = FONT_INSTRUCTION

    card_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    card_border = Border(
        left=Side(style="medium", color="D9D9D9"),
        right=Side(style="medium", color="D9D9D9"),
        top=Side(style="medium", color="D9D9D9"),
        bottom=Side(style="medium", color="D9D9D9"),
    )
    cards = [
        ("A", "B", "TOTAL INITs", len(all_inits)),
        ("C", "D", "PILARES COBERTOS", len(grouped)),
        ("E", "F", "P0", len([i for i in all_inits if i["priority"] == "P0"])),
        ("G", "H", "P1", len([i for i in all_inits if i["priority"] == "P1"])),
    ]

    row = 4
    for col_start, col_end, label, value in cards:
        ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
        c = ws.cell(row=row, column=ws[f"{col_start}1"].column, value=label)
        c.font = FONT_DASH_LABEL
        c.fill = card_fill
        c.alignment = ALIGN_CENTER
        c.border = card_border
        ws.cell(row=row, column=ws[f"{col_end}1"].column).fill = card_fill
        ws.cell(row=row, column=ws[f"{col_end}1"].column).border = card_border

        ws.merge_cells(f"{col_start}{row+1}:{col_end}{row+1}")
        c2 = ws.cell(row=row + 1, column=ws[f"{col_start}1"].column, value=value)
        c2.font = FONT_DASH_BIG
        c2.fill = card_fill
        c2.alignment = ALIGN_CENTER
        c2.border = card_border
        ws.cell(row=row + 1, column=ws[f"{col_end}1"].column).fill = card_fill
        ws.cell(row=row + 1, column=ws[f"{col_end}1"].column).border = card_border

    row = 7
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="CARTEIRA PRIORIZADA DE INITs").font = FONT_SUBTITLE

    row += 1
    headers = ["Codigo", "Titulo", "Pilar", "Subpilar", "Motor", "Prioridade", "Prazo", "EVID"]
    widths = [14, 42, 8, 10, 18, 12, 16, 22]
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for idx, init in enumerate(all_inits, start=1):
        r = row + idx
        values = [init["code"], init["title"], init["pillar"], init["subpillar"], MOTORS.get(init.get("motor", ""), init.get("motor", "")), init["priority"], init["deadline"], init.get("evid", "")]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx in (1, 3, 4, 5, 6, 7, 8) else ALIGN_WRAP

    row += len(all_inits) + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="MAPA DE PULSOS POR PILAR E SUBPILAR").font = FONT_SUBTITLE

    row += 1
    headers = ["Pilar", "Subpilar", "INITs", "Prioridade dominante", "Leitura"]
    widths = [8, 10, 10, 18, 40]
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    detail_row = row + 1
    for pillar_code in sorted(grouped.keys()):
        for subpillar_code in sorted(grouped[pillar_code].keys()):
            inits = grouped[pillar_code][subpillar_code]
            prio = next((item["priority"] for item in inits if item["priority"] == "P0"), next((item["priority"] for item in inits if item["priority"] == "P1"), "P2"))
            reading = inits[0]["title"] if inits else ""
            values = [pillar_code, subpillar_code, len(inits), prio, reading]
            for col_idx, val in enumerate(values, start=1):
                c = ws.cell(row=detail_row, column=col_idx, value=val)
                c.font = FONT_BODY
                c.border = THIN_BORDER
                c.alignment = ALIGN_CENTER if col_idx < 5 else ALIGN_WRAP
            detail_row += 1

    ws.freeze_panes = "A8"


def build_traceability_sheet(wb, area_slug):
    """Constroi aba de rastreabilidade."""
    ws = wb.create_sheet("Rastreabilidade")

    headers = ["INIT", "Titulo", "Motor", "Pilar", "Subpilar", "KR Principal", "OKR", "Tipo", "Prioridade", "Status", "Prazo", "EVID"]
    widths = [16, 45, 20, 8, 10, 14, 10, 8, 12, 16, 16, 22]

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

    ws.merge_cells("A1:L1")
    ws.cell(row=1, column=1, value="Mapa de Rastreabilidade: INIT > KR > OKR > Pilar > EVID").font = FONT_SUBTITLE

    row = 3
    for col_idx, (name, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = FONT_HEADER
        cell.fill = header_fill
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    corp_inits = [i for i in CORPORATE_INITS if area_slug in i.get("areas", [])]
    area_inits = AREA_INITS.get(area_slug, [])
    all_inits = corp_inits + area_inits
    all_inits.sort(key=lambda x: (x.get("motor", ""), x.get("priority", ""), x.get("code", "")))

    row = 4
    for init in all_inits:
        values = [
            init["code"], init["title"],
            MOTORS.get(init.get("motor", ""), ""),
            init["pillar"], init["subpillar"],
            init["kr"], init["okr"],
            init["type"], init["priority"],
            "PLANEJADA", init["deadline"],
            init.get("evid", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER if col_idx in (3, 4, 5, 7, 8, 9, 10) else ALIGN_WRAP
        row += 1

    ws.freeze_panes = "A4"


def build_corporate_scoreboard_sheet(wb):
    """Constroi o placar integrado corporativo (11-H)."""
    ws = wb.create_sheet("Placar Integrado")
    ws.sheet_properties.tabColor = PILLAR_COLORS.get("P1", {}).get("dark", "1F3864")

    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1, value="PE2026 - Placar Integrado 2026")
    cell.font = FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value="Guardrails, pilares, monetizacao e visao setorial consolidada | DOC 05 + 11-H").font = FONT_INSTRUCTION

    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="GUARDRAILS CRITICOS").font = FONT_SUBTITLE

    row += 1
    guardrail_headers = ["ID", "Foco", "Leitura 2026", "Fonte"]
    guardrails = [
        ["A1", "Margem e caixa", "Manter margem acima de 30% com caixa monitorado", "DOC 05 / 11-H"],
        ["A2", "Monetizacao", "Operar saldo, ativacao e previsibilidade 30/60/90", "DOC 05 / 11-H"],
        ["A3", "Qualidade", "Reduzir retrabalho e estabilizar SLA", "DOC 05 / 11-H"],
        ["A4", "Pessoas e governanca", "Preservar capacidade, rituais e densidade de lideranca", "DOC 05 / 11-H"],
    ]
    guardrail_widths = [10, 24, 44, 18]
    for col_idx, (header, width) in enumerate(zip(guardrail_headers, guardrail_widths), start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for idx, item in enumerate(guardrails, start=1):
        r = row + idx
        for col_idx, val in enumerate(item, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx in (1, 2, 4) else ALIGN_WRAP

    row += len(guardrails) + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="PILARES E MOTORES CORPORATIVOS").font = FONT_SUBTITLE

    row += 1
    pillar_headers = ["Pilar", "Descricao", "INITs Corp", "Motores", "Areas Afetadas"]
    pillar_widths = [8, 36, 12, 18, 34]
    for col_idx, (header, width) in enumerate(zip(pillar_headers, pillar_widths), start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for idx, pillar_code in enumerate(sorted(PILLARS.keys()), start=1):
        corp_inits = [i for i in CORPORATE_INITS if i["pillar"] == pillar_code]
        affected_areas = sorted({AREAS[a]["name"] for init in corp_inits for a in init.get("areas", [])})
        motors = sorted({init["motor"] for init in corp_inits})
        r = row + idx
        values = [pillar_code, PILLARS[pillar_code], len(corp_inits), ", ".join(motors), ", ".join(affected_areas)]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx in (1, 3, 4) else ALIGN_WRAP

    row += len(PILLARS) + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="VISAO SETORIAL CONSOLIDADA").font = FONT_SUBTITLE

    row += 1
    area_headers = ["Area", "Pilares Primarios", "INITs Corp", "INITs Setoriais", "Total", "Envelope"]
    area_widths = [24, 18, 12, 14, 10, 18]
    for col_idx, (header, width) in enumerate(zip(area_headers, area_widths), start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for idx, (slug, meta) in enumerate(sorted(AREAS.items(), key=lambda x: x[1]["code"]), start=1):
        corp_count = len([x for x in CORPORATE_INITS if slug in x.get("areas", [])])
        area_count = len(AREA_INITS.get(slug, []))
        primarios = " / ".join(AREA_PRIMARY_PILLARS.get(slug, ()))
        r = row + idx
        values = [meta["name"], primarios, corp_count, area_count, corp_count + area_count, meta["envelope"]]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx in (2, 3, 4, 5, 6) else ALIGN_LEFT

    ws.freeze_panes = "A5"


def generate_area_workbook(area_slug, output_dir):
    """Gera o workbook completo para uma area."""
    area_meta = AREAS[area_slug]
    wb = Workbook()

    setup_config_sheet(wb)
    build_dashboard_sheet(wb, area_slug, area_meta)
    build_gerencial_sheet(wb, area_slug, area_meta)
    build_direction_sheet(wb, area_slug, area_meta)
    build_action_plan_sheet(wb, area_slug, area_meta)
    build_traceability_sheet(wb, area_slug)

    wb.active = wb.sheetnames.index("Plano de Acao")

    filepath = os.path.join(output_dir, area_meta["filename"])
    wb.save(filepath)
    print(f"  [OK] {area_meta['filename']} ({area_meta['name']})")
    return filepath


def generate_corporate_workbook(output_dir):
    """Gera o workbook corporativo consolidado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Geral"

    # ---- Dashboard Geral ----
    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value="PE2026 - Visao Corporativa Consolidada").font = Font(name="Calibri", size=18, bold=True, color="1F3864")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:J2")
    ws.cell(row=2, column=1, value="Confidencial - Uso interno da Direcao | 22 INITs Corporativas | 5 Motores | Mar/2026").font = FONT_INSTRUCTION

    # Resumo por motor
    row = 4
    ws.cell(row=row, column=1, value="RESUMO POR MOTOR").font = FONT_SUBTITLE
    row = 5
    motor_headers = ["Motor", "INITs P0", "INITs P1", "Total", "Faixa Custo"]
    motor_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    for col_idx, h in enumerate(motor_headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = motor_fill
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    motor_data = [
        ("M1 - Monetizacao", 5, 3, 8, "R$ 685-785K"),
        ("M2 - Governanca", 2, 1, 3, "R$ 35-70K"),
        ("M3 - Escala", 2, 0, 2, "R$ 18-30K"),
        ("M4 - Produto/IA", 2, 2, 4, "R$ 315-515K"),
        ("M5 - Pessoas", 2, 3, 5, "R$ 385-501K"),
    ]
    for i, (motor, p0, p1, total, cost) in enumerate(motor_data):
        r = row + 1 + i
        for col_idx, val in enumerate([motor, p0, p1, total, cost], start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx > 1 else ALIGN_LEFT

    # Resumo por area
    row = 12
    ws.cell(row=row, column=1, value="DISTRIBUICAO POR AREA").font = FONT_SUBTITLE
    row = 13
    area_headers = ["Area", "INITs Corp.", "INITs Setoriais", "Total", "Envelope"]
    for col_idx, h in enumerate(area_headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = motor_fill
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, (slug, meta) in enumerate(sorted(AREAS.items(), key=lambda x: x[1]["code"])):
        r = row + 1 + i
        corp_count = len([x for x in CORPORATE_INITS if slug in x.get("areas", [])])
        area_count = len(AREA_INITS.get(slug, []))
        for col_idx, val in enumerate([meta["name"], corp_count, area_count, corp_count + area_count, meta["envelope"]], start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx > 1 else ALIGN_LEFT

    for col in "ABCDE":
        ws.column_dimensions[col].width = 22

    # ---- Placar Integrado ----
    build_corporate_scoreboard_sheet(wb)

    # ---- Carteira P0 ----
    ws_p0 = wb.create_sheet("Carteira P0")
    setup_corporate_carteira(ws_p0, "P0", [i for i in CORPORATE_INITS if i["priority"] == "P0"])

    # ---- Carteira P1 ----
    ws_p1 = wb.create_sheet("Carteira P1")
    setup_corporate_carteira(ws_p1, "P1", [i for i in CORPORATE_INITS if i["priority"] == "P1"])

    # ---- Por Area ----
    ws_area = wb.create_sheet("Por Area")
    setup_corporate_by_area(ws_area)

    # ---- Rastreabilidade ----
    ws_trace = wb.create_sheet("Rastreabilidade")
    setup_corporate_traceability(ws_trace)

    # Config
    setup_config_sheet(wb)

    wb.active = wb.sheetnames.index("Placar Integrado")

    filepath = os.path.join(output_dir, "PE2026_Corporativo.xlsx")
    wb.save(filepath)
    print(f"  [OK] PE2026_Corporativo.xlsx (Consolidado)")
    return filepath


def setup_corporate_carteira(ws, priority_label, inits):
    """Setup a carteira sheet (P0 or P1)."""
    ws.merge_cells("A1:Q1")
    ws.cell(row=1, column=1,
            value=f"PE2026 - Carteira {priority_label} ({'Critico 90 dias' if priority_label == 'P0' else 'Estrategico do ano'})").font = FONT_TITLE

    ws.merge_cells("A2:Q2")
    ws.cell(row=2, column=1,
            value=f"{len(inits)} iniciativas | Agrupadas por Motor | DOC 08 v2").font = FONT_INSTRUCTION

    row = 4
    write_header_row(ws, row)

    current_row = 5
    current_motor = None
    inits_sorted = sorted(inits, key=lambda x: (x["motor"], x["code"]))

    for init in inits_sorted:
        if init["motor"] != current_motor:
            current_motor = init["motor"]
            motor_label = MOTORS.get(current_motor, current_motor)
            fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(COLUMNS))
            c = ws.cell(row=current_row, column=1, value=motor_label)
            c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            c.fill = fill
            for col in range(2, len(COLUMNS) + 1):
                ws.cell(row=current_row, column=col).fill = fill
            current_row += 1

        write_init_row(ws, current_row, init)
        current_row += 1

    ws.freeze_panes = "A5"


def setup_corporate_by_area(ws):
    """Tabela cruzada INIT x Area."""
    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value="PE2026 - Distribuicao de INITs por Area").font = FONT_TITLE

    headers = ["INIT", "Titulo", "Motor", "Prioridade", "RH", "MKT", "P&D", "Operacao", "CS", "Comercial", "Financeiro"]
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

    row = 3
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = header_fill
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    area_order = ["rh", "marketing", "pd", "operacoes", "cs", "comercial", "financeiro"]

    row = 4
    for init in sorted(CORPORATE_INITS, key=lambda x: x["code"]):
        values = [init["code"], init["title"], MOTORS.get(init["motor"], ""), init["priority"]]
        for area_slug in area_order:
            values.append("X" if area_slug in init.get("areas", []) else "")
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx > 2 else ALIGN_WRAP
            if val == "X":
                c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                c.font = Font(name="Calibri", size=11, bold=True, color="375623")
        row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    for col in "EFGHIJK":
        ws.column_dimensions[col].width = 12

    ws.freeze_panes = "A4"


def setup_corporate_traceability(ws):
    """Rastreabilidade completa de todas as 22 INITs."""
    ws.merge_cells("A1:L1")
    ws.cell(row=1, column=1, value="PE2026 - Mapa de Rastreabilidade Integrado").font = FONT_TITLE

    headers = ["INIT", "Titulo", "Motor", "Pilar", "Subpilar", "OKR", "KR Principal", "Tipo", "Prioridade", "Area(s)", "Custo (R$)", "EVID"]
    widths = [14, 45, 20, 8, 10, 10, 14, 8, 12, 22, 16, 22]

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

    row = 3
    for col_idx, (name, width) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=col_idx, value=name)
        c.font = FONT_HEADER
        c.fill = header_fill
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row = 4
    for init in sorted(CORPORATE_INITS, key=lambda x: (x["motor"], x["priority"], x["code"])):
        area_names = ", ".join([AREAS.get(a, {}).get("name", a) for a in init.get("areas", [])])
        values = [
            init["code"], init["title"],
            MOTORS.get(init["motor"], ""), init["pillar"], init["subpillar"],
            init["okr"], init["kr"], init["type"], init["priority"],
            area_names, init["cost"], init.get("evid", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if col_idx in (3, 4, 5, 6, 8, 9) else ALIGN_WRAP
        row += 1

    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────

def main():
    output_root = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                               "docs", "planos-acao")
    output_dir = os.path.join(output_root, "pa.2026")
    os.makedirs(output_dir, exist_ok=True)

    print("=" * 60)
    print("PE2026 - Gerador de Planos de Acao Setoriais")
    print("=" * 60)
    print(f"Destino: {output_dir}\n")

    print("Gerando 7 workbooks setoriais...")
    for area_slug in ["operacoes", "cs", "rh", "marketing", "pd", "financeiro", "comercial"]:
        generate_area_workbook(area_slug, output_dir)

    print("\nGerando workbook corporativo...")
    generate_corporate_workbook(output_dir)

    print(f"\n{'=' * 60}")
    print(f"CONCLUIDO: 8 workbooks gerados em {output_dir}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
